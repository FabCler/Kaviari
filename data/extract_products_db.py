#!/usr/bin/env python3
"""Build the v2 product catalog from Data_base_products.xlsx (the canonical
product database: PR code, description, caviar type, category, unit,
packing/box) and join it with the Import Review workbook for stock on hand
and the real weekly consumption history (matched by PR code).

Usage: python3 data/extract_products_db.py

Outputs (next to this script):
  - products_db.json        catalog for prisma/seed.ts
  - consumption_history.json  per-prCode weekly consumption (units/week)

Unit costs are estimated (no prices in either source) — "costIsEstimate".
"""
import json
import re
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")
OUT = Path(__file__).parent

WEEKLY_SHEETS = [
    "17.06.26", "23.06.26", "30.06.26", "07.07.26",
    "14.07.26", "21.07.26", "28.07.26", "04.08.26", "10.08.26",
]

EUR_PER_KG = {
    "Beluga": 3400, "Daurikus": 1500, "Oscietra": 1250, "Kristal": 1050,
    "Transmontanus": 1000, "Baeri": 800, "Baenki": 850,
}
# Non-caviar cost estimates (EUR per unit).
OTHER_COST = {
    "Fish Roe": 12.0, "Fish": 60.0, "Crab": 85.0, "Marketing Tools": 5.0,
}

GRAMS_RE = re.compile(r"\((\d+)\s*g(?:r|m)?\s*/\s*tin\)|(\d+)\s*g(?:r|m)?\s*/\s*tin|-\s*(\d+)\s*gr\b", re.IGNORECASE)


def grams_from_description(desc: str):
    m = GRAMS_RE.search(desc)
    if not m:
        return None
    return int(next(g for g in m.groups() if g))


def main() -> None:
    db = openpyxl.load_workbook(OUT / "Data_base_products.xlsx", data_only=True)
    rows = [
        r for r in db.worksheets[0].iter_rows(min_row=4, values_only=True)
        if r[0] is not None
    ]

    # ---- stock + weekly history from the import-review workbook, by PR code
    stock: dict[str, float] = {}
    history: dict[str, list[float]] = defaultdict(lambda: [0.0] * len(WEEKLY_SHEETS))
    review_path = OUT / "Import_Review_Kaviari.xlsx"
    if review_path.exists():
        review = openpyxl.load_workbook(review_path, data_only=True)
        sheets = {ws.title: ws for ws in review.worksheets}
        for idx, name in enumerate(WEEKLY_SHEETS):
            ws = sheets[name]
            for r in ws.iter_rows(min_row=3, values_only=True):
                if not r or r[0] is None:
                    continue
                pr = str(r[0])
                cons = float(r[18] or 0)
                # A PR code can appear once per sheet; last one wins.
                if pr not in history and cons == 0 and name != WEEKLY_SHEETS[-1]:
                    pass
                history[pr][idx] = cons
                if name == WEEKLY_SHEETS[-1]:
                    stock[pr] = float(r[16] or 0)

    catalog = []
    seen = set()
    for pr_code, desc, caviar_type, category, unit, packing, _pack_unit in rows:
        pr = str(pr_code)
        if pr in seen:
            continue
        seen.add(pr)
        desc = str(desc).strip()
        grams = grams_from_description(desc) if unit == "Tin" else None
        if category == "Caviar":
            per_kg = EUR_PER_KG.get(str(caviar_type), 1250)
            cost = round(per_kg * (grams or 30) / 1000, 2)
        else:
            cost = OTHER_COST.get(str(category), 10.0)
        catalog.append({
            "prCode": pr,
            "name": desc,
            "caviarType": str(caviar_type) if caviar_type else None,
            "category": str(category),
            "unit": str(unit),
            "packingPerBox": int(packing) if packing else None,
            "gramsPerUnit": grams,
            "unitCost": cost,
            "costIsEstimate": True,
            "currency": "EUR",
            "stockOnHand": stock.get(pr, 0.0),
        })

    (OUT / "products_db.json").write_text(
        json.dumps(catalog, indent=2, ensure_ascii=False)
    )
    hist_out = {
        "weeks": WEEKLY_SHEETS,
        "unit": "units/week",
        "consumption": {p["prCode"]: history[p["prCode"]] for p in catalog if p["prCode"] in history},
    }
    (OUT / "consumption_history.json").write_text(
        json.dumps(hist_out, indent=2, ensure_ascii=False)
    )

    movers = sum(1 for p in catalog if sum(history.get(p["prCode"], [])) > 0)
    print(f"{len(catalog)} products, {movers} with movement, "
          f"{sum(1 for p in catalog if p['stockOnHand'] > 0)} with stock")


if __name__ == "__main__":
    main()
